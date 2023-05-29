import inspect
import logging

from openpyxl import Workbook, load_workbook

# def custom_logger(logLevel=logging.DEBUG):
#     logger_name = inspect.stack()[1][3]
#     logger = logging.getLogger(logger_name)
#     logger.setLevel(logLevel)
#     fh = logging.FileHandler("automation.log")
#     formatter = logging.Formatter('%(acstime)s: %(levelname)s: %(message)s,', datefmt='%m/%d/%Y %I:%M:%S %p')
#     fh.setFormatter(formatter)
#     logger.addHandler(fh)
#     return logger

import inspect
import logging


class log_class:
    @staticmethod
    def custom_logger():
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(
            loggerName)  # it will tell that right now this particular test file is being run and the logs are printed in regarding to that test file
        filename = logging.FileHandler(
            'logfile.log')  # fileHandler method gives info. about the logging file where all the logs will be printed
        formatter = logging.Formatter(
            "%(asctime)s : %(levelname)s : %(name)s : %(message)s", datefmt="%m/%d/%Y %I:%M:%S %p")  # here the format is decided like in what format the log message is printed
        filename.setFormatter(
            formatter)  # here the format is appended to filehandler to pass the format info. to log file

        logger.addHandler(filename)

        logger.setLevel(
            logging.INFO)  # it will now print all the messages from and below the info level and skip the debug messages
        # below is the order of the type of logs
        # logger.debug("This is a debug statement")
        # logger.info("This is information")
        # logger.warning("This is a warning")
        # logger.error("This is a error")
        # logger.critical("This is a critical information")
        return logger


class Utils:
    def read_data_from_excel(file_name, sheet):
        dataList = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            dataList.append(row)
        return dataList
